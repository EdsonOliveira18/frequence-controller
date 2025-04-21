import openpyxl
from datetime import datetime, timedelta
import os

ARQUIVO_PLANILHA = "presencas.xlsx"
cache_registros = {}
ultima_atualizacao = datetime.min

def inicializar_planilha():
    if not os.path.exists(ARQUIVO_PLANILHA):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Presenças"
        ws.append(["CPF", "Data", "Horário Início 1", "Horário Término 1", "Horário Início 2", "Horário Término 2"])
        wb.save(ARQUIVO_PLANILHA)

def carregar_cache():
    global cache_registros, ultima_atualizacao
    wb = openpyxl.load_workbook(ARQUIVO_PLANILHA)
    ws = wb.active
    cache_registros = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cpf, data = row[0], row[1]
        chave = (cpf, data)
        horarios = row[2:]
        cache_registros[chave] = horarios
    wb.close()
    ultima_atualizacao = datetime.now()

def salvar_cache_no_excel():
    global cache_registros
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Presenças"
    ws.append(["CPF", "Data", "Horário Início 1", "Horário Término 1", "Horário Início 2", "Horário Término 2"])
    for (cpf, data), horarios in cache_registros.items():
        ws.append([cpf, data] + list(horarios))
    wb.save(ARQUIVO_PLANILHA)
    wb.close()

def registrar_presenca(data):
    inicializar_planilha()
    agora = datetime.now()

    # Atualiza cache se tiver mais de 1 minuto desde última atualização
    if (agora - ultima_atualizacao) > timedelta(minutes=1):
        carregar_cache()

    cpf = data["cpf"]
    horario = data["horario"]
    data_hoje = agora.strftime("%d/%m/%Y")
    chave = (cpf, data_hoje)

    if chave not in cache_registros:
        cache_registros[chave] = [None, None, None, None]

    horarios = cache_registros[chave]

    # Evita duplicação de horário
    if horario in horarios:
        return "Este horário já foi registrado."

    for i in range(4):
        if horarios[i] is None:
            horarios[i] = horario
            break
    else:
        return "Limite de horários atingido para hoje."

    # Salva no Excel apenas a cada minuto
    salvar_cache_no_excel()
    return "Presença registrada com sucesso."
